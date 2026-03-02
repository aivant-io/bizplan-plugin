#!/usr/bin/env python3
"""
Validate a populated ecommerce financial model Excel file.

Checks:
  1. All 49 input cells are populated (not empty/zero where unexpected)
  2. Type correctness (percentages as decimals, years as integers)
  3. Balance sheet check (Assets - Liabilities - Equity ~ 0)
  4. Traffic mix sums to 1.0
  5. Year range validation (2026-2031)
  6. Conversion rate hierarchy (retention >= organic >= paid)

Usage:
    python validate_model.py <path_to_populated_model.xlsx>
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl>=3.1.2")
    sys.exit(1)


# All 49 input cells on the Assumptions sheet
INPUT_CELLS = {
    "B5": ("traffic_total_year1", "int"),
    "B6": ("traffic_yoy_growth", "float_pct"),
    "B9": ("traffic_mix_paid_pct", "float_pct"),
    "B10": ("traffic_mix_organic_pct", "float_pct"),
    "B11": ("traffic_mix_retention_pct", "float_pct"),
    "B15": ("conv_paid", "float_pct"),
    "B16": ("conv_organic", "float_pct"),
    "B17": ("conv_retention", "float_pct"),
    "B20": ("repeat_purchase_rate_year1", "float_pct"),
    "B21": ("repeat_purchase_frequency", "float"),
    "B22": ("retention_improvement_annual", "float_pct"),
    "B25": ("aov_year1", "currency"),
    "B26": ("aov_inflation", "float_pct"),
    "B27": ("cogs_pct", "float_pct"),
    "B28": ("cogs_annual_improvement_pct", "float_pct"),
    "B29": ("discounts_promos_pct", "float_pct"),
    "B30": ("return_rate_pct", "float_pct"),
    "B31": ("payment_processing_pct", "float_pct"),
    "B32": ("platform_fee_pct", "float_pct"),
    "B35": ("cpc_year1", "currency"),
    "B36": ("cpc_inflation", "float_pct"),
    "B39": ("ship_cost_year1", "currency"),
    "B40": ("ship_inflation", "float_pct"),
    "B41": ("handling_cost_year1", "currency"),
    "B42": ("packaging_cost_per_order", "currency"),
    "B43": ("support_cost_per_order", "currency"),
    "B46": ("warehouse_rent_amount_when_active", "currency"),
    "B47": ("warehouse_rent_start_year", "year"),
    "B48": ("warehouse_rent_inflation", "float_pct"),
    "B51": ("office_rent_amount_when_active", "currency"),
    "B52": ("office_rent_start_year", "year"),
    "B53": ("office_rent_inflation", "float_pct"),
    "B54": ("salaries_year1", "currency"),
    "B55": ("months_to_full_team", "int"),
    "B56": ("prof_fees_year1", "currency"),
    "B57": ("sal_prof_inflation", "float_pct"),
    "B58": ("misc_ga_year1", "currency"),
    "B59": ("monthly_saas_costs", "currency"),
    "B62": ("ar_days", "float"),
    "B63": ("inventory_days_year1", "float"),
    "B64": ("inventory_turns_improvement", "float_pct"),
    "B65": ("ap_days", "float"),
    "B68": ("capex_pct_of_net_rev", "float_pct"),
    "B69": ("dep_period_years", "int"),
    "B72": ("initial_loan_amount", "currency"),
    "B74": ("interest_rate", "float_pct"),
    "B75": ("loan_term_years", "int"),
    "B76": ("initial_equity_injection", "currency"),
    "B79": ("tax_rate", "float_pct"),
}

# Balance check cells on Model sheet
BALANCE_CHECK_CELLS = {
    "year_1": "E194",
    "year_2": "F194",
    "year_3": "G194",
    "year_4": "H194",
    "year_5": "I194",
    "year_6": "J194",
}

# Percentage bounds (should be <= 1.0 if stored as decimal)
PCT_MAX_REASONABLE = 1.0


def repair_xlsx_calc_settings(xlsx_path: str) -> list[str]:
    """Auto-fix fullCalcOnLoad and calcChain in the xlsx ZIP. Returns list of repairs made."""
    import zipfile as _zf

    repairs = []
    path = Path(xlsx_path)

    with _zf.ZipFile(str(path), "r") as zf:
        files = {name: zf.read(name) for name in zf.namelist()}

    # Fix 1: Ensure fullCalcOnLoad="1" on <calcPr>
    if "xl/workbook.xml" in files:
        wb_xml = files["xl/workbook.xml"].decode("utf-8")
        if 'fullCalcOnLoad="1"' not in wb_xml:
            wb_xml = re.sub(r'\s*fullCalcOnLoad="[^"]*"', '', wb_xml)
            wb_xml = re.sub(r'<calcPr\b', r'<calcPr fullCalcOnLoad="1"', wb_xml)
            files["xl/workbook.xml"] = wb_xml.encode("utf-8")
            repairs.append('AUTO-FIX: Added fullCalcOnLoad="1" to workbook.xml')

    # Fix 2: Remove calcChain.xml if present
    if "xl/calcChain.xml" in files:
        del files["xl/calcChain.xml"]
        if "[Content_Types].xml" in files:
            ct = files["[Content_Types].xml"].decode("utf-8")
            ct = re.sub(r'<Override[^>]*calcChain[^>]*/>', '', ct)
            files["[Content_Types].xml"] = ct.encode("utf-8")
        repairs.append("AUTO-FIX: Removed calcChain.xml from ZIP")

    # Write back only if repairs were made
    if repairs:
        with _zf.ZipFile(str(path), "w", _zf.ZIP_DEFLATED) as zf:
            for name, content in files.items():
                zf.writestr(name, content)

    return repairs


def validate_model(xlsx_path: str) -> tuple[bool, list[str], list[str]]:
    """
    Validate a populated financial model.

    Returns:
        (passed, errors, warnings)
    """
    errors: list[str] = []
    warnings: list[str] = []

    path = Path(xlsx_path)
    if not path.exists():
        return False, [f"File not found: {xlsx_path}"], []

    wb = openpyxl.load_workbook(str(path), data_only=True)

    # --- Check 1: Assumptions sheet exists ---
    if "Assumptions" not in wb.sheetnames:
        return False, ["'Assumptions' sheet not found in workbook"], []

    ws = wb["Assumptions"]

    # --- Check 2: All 49 input cells populated ---
    populated_count = 0
    for cell_addr, (input_id, value_type) in INPUT_CELLS.items():
        value = ws[cell_addr].value
        if value is None:
            errors.append(f"MISSING: {cell_addr} ({input_id}) is empty")
        else:
            populated_count += 1

            # --- Check 3: Type correctness ---
            if value_type == "float_pct":
                if isinstance(value, (int, float)) and value > PCT_MAX_REASONABLE:
                    errors.append(
                        f"PCT_ERROR: {cell_addr} ({input_id}) = {value} "
                        f"(should be decimal, e.g., 0.15 for 15%)"
                    )
            elif value_type == "int":
                if isinstance(value, float) and value != int(value):
                    warnings.append(
                        f"TYPE_WARN: {cell_addr} ({input_id}) = {value} "
                        f"(expected integer)"
                    )
            elif value_type == "year":
                if isinstance(value, (int, float)):
                    year_val = int(value)
                    if year_val < 2026 or year_val > 2031:
                        errors.append(
                            f"YEAR_ERROR: {cell_addr} ({input_id}) = {year_val} "
                            f"(must be 2026-2031)"
                        )
                else:
                    errors.append(
                        f"TYPE_ERROR: {cell_addr} ({input_id}) = {value} "
                        f"(expected year integer)"
                    )

    # --- Check 4: Traffic mix sums to 1.0 ---
    paid = ws["B9"].value or 0
    organic = ws["B10"].value or 0
    retention = ws["B11"].value or 0
    mix_sum = paid + organic + retention
    if abs(mix_sum - 1.0) > 0.01:
        errors.append(
            f"MIX_ERROR: Traffic mix sum = {mix_sum:.4f} "
            f"(paid={paid} + organic={organic} + retention={retention}), "
            f"should be 1.0"
        )

    # --- Check 5: Conversion hierarchy ---
    conv_paid = ws["B15"].value or 0
    conv_organic = ws["B16"].value or 0
    conv_retention = ws["B17"].value or 0
    if conv_paid > conv_organic:
        warnings.append(
            f"CONV_WARN: conv_paid ({conv_paid}) > conv_organic ({conv_organic})"
        )
    if conv_organic > conv_retention:
        warnings.append(
            f"CONV_WARN: conv_organic ({conv_organic}) > conv_retention ({conv_retention})"
        )

    # --- Check 6: Balance sheet check ---
    if "Model" in wb.sheetnames:
        model_ws = wb["Model"]
        none_count = 0
        for year_label, cell_addr in BALANCE_CHECK_CELLS.items():
            check_val = model_ws[cell_addr].value
            if check_val is None:
                none_count += 1
            elif abs(check_val) > 0.01:
                errors.append(
                    f"BALANCE_ERROR: {year_label} balance check = {check_val:.4f} "
                    f"(should be ~0)"
                )
        if none_count == len(BALANCE_CHECK_CELLS):
            warnings.append(
                "BALANCE_WARN: All balance check cells (E194:J194) returned None. "
                "Formula cached values were cleared — cannot verify balance sheet "
                "from data_only mode. Open the file in Excel and confirm formulas "
                "recalculate correctly."
            )
        elif none_count > 0:
            warnings.append(
                f"BALANCE_WARN: {none_count}/{len(BALANCE_CHECK_CELLS)} balance "
                f"check cells returned None — partial verification only."
            )
    else:
        warnings.append("'Model' sheet not found — skipping balance check")

    wb.close()

    # --- Check 7: Auto-fix XML-level calc settings ---
    try:
        repairs = repair_xlsx_calc_settings(str(path))
        warnings.extend(repairs)
    except Exception as e:
        warnings.append(f"REPAIR_WARN: Could not inspect/repair ZIP: {e}")

    passed = len(errors) == 0
    return passed, errors, warnings


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <path_to_model.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    print(f"Validating: {xlsx_path}")
    print("=" * 60)

    passed, errors, warnings = validate_model(xlsx_path)

    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors:
            print(f"  {e}")

    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  {w}")

    print(f"\n{'PASSED' if passed else 'FAILED'}")
    print(f"  Total input cells checked: {len(INPUT_CELLS)}")
    print(f"  Errors: {len(errors)}")
    print(f"  Warnings: {len(warnings)}")

    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
